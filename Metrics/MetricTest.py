import os
import pprint
import time

import numpy as np
import openpyxl
import torch
from Modules.Loss import (BendingEnergyMetric, DiceCoefficient,
                          JacobianDeterminantMetric, SurfaceDistanceFromSeg)
from openpyxl import Workbook, worksheet
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


class MetricTest:
    def __init__(self):
        self.details = {
            'York': {},
            'ACDC': {},
            'MICCAI': {},
            'M&M': {},
        }
        self.info = {
            'York': {
                'LvMyo': [1],
                'LvBp': [2],
                'Lv': [1, 2]
            },
            'ACDC': {
                'Rv': [1],
                'LvMyo': [2],
                'LvBp': [3],
                'Lv': [2, 3],
                # 'LvRv': [1, 2, 3]
            },
            'MICCAI': {
                'LvBp': [1]
            },
            'M&M': {
                'LvBp': [1],
                'LvMyo': [2],
                'Rv': [3],
                'Lv': [1, 2]
            },
        }
        self.dice_estimate = DiceCoefficient()
        self.surface_dist_estimate = SurfaceDistanceFromSeg()
        self.be_estimate = BendingEnergyMetric()
        self.jacobian_estiamte = JacobianDeterminantMetric()

    def getDatasetName(self, case_no):
        if case_no <= 33:
            return 'York'
        elif case_no > 33 and case_no <= 78:
            return 'MICCAI'
        elif case_no > 78 and case_no <= 228:
            return 'ACDC'
        else:
            return 'M&M'

    def testMetrics(self, src, wraped_src, tgt, wraped_tgt, resolution,
                    case_no, slc_idx):
        dataset_name = self.getDatasetName(case_no)
        self.details[dataset_name][case_no] = {}
        self.details[dataset_name][case_no]['slc_idx'] = [
            t.item() for t in slc_idx
        ]

        for key_name in self.info[dataset_name]:
            selected_src = torch.zeros_like(src)
            selected_wraped_src = torch.zeros_like(wraped_src)
            selected_tgt = torch.zeros_like(tgt)
            selected_wraped_tgt = torch.zeros_like(wraped_tgt)
            for v in self.info[dataset_name][key_name]:
                selected_src += (src == v)
                selected_wraped_src += (wraped_src == v)
                selected_tgt += (tgt == v)
                selected_wraped_tgt += (wraped_tgt == v)
            self.details[dataset_name][case_no][key_name] = {
                'undef':
                self.testOnePair(selected_src, selected_tgt, resolution),
                'ed_to_es':
                self.testOnePair(selected_wraped_src, selected_tgt,
                                 resolution),
                'es_to_ed':
                self.testOnePair(selected_wraped_tgt, selected_src, resolution)
            }

    def testOnePair(self, seg_pred_batch, seg_gt_batch, resolution):
        dice_result = self.dice_estimate(seg_pred_batch,
                                         seg_gt_batch).cpu().numpy()
        dice_result = dice_result[np.logical_not(np.isnan(dice_result))]
        # dice_result = dice_result[np.logical_not(dice_result==0)]
        seg_gt_batch = seg_gt_batch.cpu().numpy()[:, 0]
        seg_pred_batch = seg_pred_batch.cpu().numpy()[:, 0]
        apd_result = []
        hd_result = []
        APDgt_pred_result = []
        APDpred_gt_result = []
        for seg_gt, seg_pred in zip(seg_gt_batch, seg_pred_batch):
            # print(seg_gt.shape)
            surface_dist = self.surface_dist_estimate.compute_surface_distances(
                seg_gt, seg_pred, resolution)
            apd = self.surface_dist_estimate.compute_average_surface_distance(
                surface_dist)
            sym_apd = (apd[0] + apd[1]) / 2
            hd = self.surface_dist_estimate.compute_robust_hausdorff(
                surface_dist, 95)
            if not np.isposinf(sym_apd) and not np.isposinf(hd):
                apd_result.append(sym_apd)
                hd_result.append(hd)
                APDgt_pred_result.append(apd[0])
                APDpred_gt_result.append(apd[1])

        return {
            'Dice': dice_result,
            'SymAPD': apd_result,
            'HD': hd_result,
            'APDgt_pred': APDgt_pred_result,
            'APDpred_gt': APDpred_gt_result
        }

    def testFlow(self, flows_t, flowt_s, case_no):
        dataset_name = self.getDatasetName(case_no)
        bes_t_result = self.be_estimate(flows_t).cpu().numpy()
        jacobians_t_result = torch.sum(self.jacobian_estiamte(flows_t) <= 0,
                                       dim=[1, 2]).cpu().numpy()
        bet_s_result = self.be_estimate(flowt_s).cpu().numpy()
        jacobiant_s_result = torch.sum(self.jacobian_estiamte(flowt_s) <= 0,
                                       dim=[1, 2]).cpu().numpy()
        self.details[dataset_name][case_no]['flow'] = {
            'ed_to_es': {
                'BE': bes_t_result * 1e4,
                'Jacobian': jacobians_t_result
            },
            'es_to_ed': {
                'BE': bet_s_result * 1e4,
                'Jacobian': jacobiant_s_result
            }
        }

    def meanByAnotomicalOfDataset(self):
        # average along with the same anotomical of dataset
        # first average in the same case
        self.mean_by_anotomical = {}

        for dataset_name in self.details:
            dataset = self.details[dataset_name]
            self.mean_by_anotomical[dataset_name] = {}
            for anotomical in self.info[dataset_name]:
                self.mean_by_anotomical[dataset_name][anotomical] = {}
                for direction in ['undef', 'ed_to_es', 'es_to_ed']:
                    self.mean_by_anotomical[dataset_name][anotomical][
                        direction] = {}
                    for metric_name in [
                            'Dice', 'SymAPD', 'HD', 'APDgt_pred', 'APDpred_gt'
                    ]:
                        case_mean = [
                            np.mean(dataset[case_no][anotomical][direction]
                                    [metric_name]) for case_no in dataset
                        ]
                        self.mean_by_anotomical[dataset_name][anotomical][
                            direction][metric_name] = {
                                'mean': np.mean(case_mean),
                                'std': np.std(case_mean)
                            }

    def meanByDataset(self):
        # first average in a case
        # then average along with cases

        self.mean_by_dataset = {}

        for dataset_name in self.details:
            dataset = self.details[dataset_name]
            self.mean_by_dataset[dataset_name] = {}
            for direction in ['undef', 'ed_to_es', 'es_to_ed']:
                self.mean_by_dataset[dataset_name][direction] = {}
                for metric_name in [
                        'Dice', 'SymAPD', 'HD', 'APDgt_pred', 'APDpred_gt'
                ]:
                    case_mean = []
                    for case_no in dataset:
                        case = dataset[case_no]
                        case_item = np.array([
                            case[anotomical][direction][metric_name]
                            for anotomical in self.info[dataset_name]
                        ])
                        case_mean.append(
                            np.mean(np.mean(case_item, axis=0), axis=0))
                    self.mean_by_dataset[dataset_name][direction][
                        metric_name] = {
                            'mean': np.mean(case_mean),
                            'std': np.std(case_mean)
                        }
            # flow
            for direction in ['ed_to_es', 'es_to_ed']:
                for metric_name in ['BE', 'Jacobian']:
                    case_mean = [
                        np.mean(
                            dataset[case_no]['flow'][direction][metric_name])
                        for case_no in dataset
                    ]
                    self.mean_by_dataset[dataset_name][direction][
                        metric_name] = {
                            'mean': np.mean(case_mean),
                            'std': np.std(case_mean)
                        }

    def meanByAll(self):
        # 1. average along with anotomical in a slice
        # 2. average along with slice
        # 3. average along with case_no
        self.mean_by_all = {}

        for direction in ['undef', 'ed_to_es', 'es_to_ed']:
            self.mean_by_all[direction] = {}
            for metric_name in [
                    'Dice', 'SymAPD', 'HD', 'APDgt_pred', 'APDpred_gt'
            ]:
                case_mean = []
                for dataset_name in self.details:
                    dataset = self.details[dataset_name]
                    for case_no in dataset:
                        case = dataset[case_no]
                        case_item = np.array([
                            case[anotomical][direction][metric_name]
                            for anotomical in self.info[dataset_name]
                        ])
                        case_mean.append(
                            np.mean(np.mean(case_item, axis=0), axis=0))
                self.mean_by_all[direction][metric_name] = {
                    'mean': np.mean(case_mean),
                    'std': np.std(case_mean)
                }

        # flow
        for direction in ['ed_to_es', 'es_to_ed']:
            for metric_name in ['BE', 'Jacobian']:
                case_mean = []
                for dataset_name in self.details:
                    dataset = self.details[dataset_name]
                    case_mean += [
                        np.mean(
                            dataset[case_no]['flow'][direction][metric_name])
                        for case_no in dataset
                    ]
                self.mean_by_all[direction][metric_name] = {
                    'mean': np.mean(case_mean),
                    'std': np.std(case_mean)
                }

    def mean(self):
        self.meanByAnotomicalOfDataset()
        self.meanByDataset()
        self.meanByAll()
        return self.mean_by_all['ed_to_es']['Dice']

    def cellValue(self, ws, row, column, value):
        alignment = Alignment(vertical='center', horizontal='center')
        cell = ws.cell(row=row, column=column)
        cell.alignment = alignment
        cell.value = value

    def getWidth(self, name: str):
        width = 0
        for c in name:
            if c.isdigit():
                width += 1
            elif c.isupper():
                width += 1.2
            else:
                width += 1.1

        return width

    def appendData(self, wb_save_path, network, name, start_row):
        direction = "ed_to_es"
        wb = openpyxl.load_workbook(wb_save_path)
        ws = wb[direction]
        alignment = Alignment(vertical='center', horizontal='center')
        self.cellValue(ws, start_row, 1, network)
        self.cellValue(ws, start_row, 2, name)
        metric_name = [
            'BE', 'Jacobian', 'Dice', 'HD', 'SymAPD', 'APDgt_pred',
            'APDpred_gt'
        ]
        metric_format = [
            '%.2f(%.2f)', '%.2f(%.2f)', '%.3f(%.3f)', '%.2f(%.2f)',
            '%.2f(%.2f)', '%.2f(%.2f)', '%.2f(%.2f)'
        ]
        for i in range(3, 10):
            if metric_name[i - 3] in self.mean_by_all[direction]:
                metric_value = self.mean_by_all[direction][metric_name[i - 3]]
                value = metric_format[i - 3] % (metric_value['mean'],
                                                metric_value['std'])
                self.cellValue(ws, start_row, i, value)
        start = 10
        for dataset_name in ['York', 'ACDC', 'MICCAI','M&M']:
            for i in range(7):
                if metric_name[i] in self.mean_by_dataset[dataset_name][
                        direction]:
                    metric_value = self.mean_by_dataset[dataset_name][
                        direction][metric_name[i]]
                    value = metric_format[i] % (metric_value['mean'],
                                                metric_value['std'])
                    self.cellValue(ws, start_row, i + start, value)
            start += 7

        metric_name = ['Dice', 'HD', 'SymAPD', 'APDgt_pred', 'APDpred_gt']
        metric_format = [
            '%.3f(%.3f)', '%.2f(%.2f)', '%.2f(%.2f)', '%.2f(%.2f)',
            '%.2f(%.2f)'
        ]
        
        for dataset_name in ['York', 'ACDC', 'MICCAI','M&M']:
            anatomical_list = list(self.info[dataset_name])
            for anatomical_name in anatomical_list:
                for i in range(5):
                    metric_value = self.mean_by_anotomical[dataset_name][
                        anatomical_name][direction][metric_name[i]]
                    value = metric_format[i] % (metric_value['mean'],
                                                metric_value['std'])
                    self.cellValue(ws, start_row, i + start, value)
                start += 5
        
        wb.save(wb_save_path)


    def autoSetWidth(self, ws):
        col_width = []
        #获取每一列的内容的最大宽度
        i = 0
        # 每列
        for col in ws.columns:
            # 每行
            for j in range(len(col)):
                if j == 0:
                    # 数组增加一个元素
                    col_width.append(self.getWidth(str(col[j].value)))
                else:
                    # 获得每列中的内容的最大宽度
                    if col_width[i] < self.getWidth(str(col[j].value)):
                        col_width[i] = self.getWidth(str(col[j].value))
            i = i + 1

        #设置列宽
        for i in range(len(col_width)):
            # 根据列的数字返回字母
            col_letter = get_column_letter(i + 1)
            # 当宽度大于100，宽度设置为100
            if col_width[i] > 100:
                ws.column_dimensions[col_letter].width = 100
            # 只有当宽度大于10，才设置列宽
            elif col_width[i] > 10:
                ws.column_dimensions[col_letter].width = col_width[i] + 2

    def initWorksheet(self, ws: worksheet.worksheet.Worksheet):
        alignment = Alignment(vertical='center', horizontal='center')
        ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
        self.cellValue(ws, 1, 1, 'Network')
        ws.merge_cells(start_row=1, start_column=2, end_row=3, end_column=2)
        self.cellValue(ws, 1, 2, 'Name')
        metric_name = [
            'BE', 'Jacobian', 'Dice', 'HD', 'SymAPD', 'APDgt_pred',
            'APDpred_gt'
        ]
        start = 3
        for dataset_name in ['all', 'York', 'ACDC', 'MICCAI','M&M']:
            ws.merge_cells(start_row=1,
                           start_column=start,
                           end_row=2,
                           end_column=start + 6)
            self.cellValue(ws, 1, start, dataset_name)
            for i in range(7):
                self.cellValue(ws, 3, start + i, metric_name[i])
            start += 7
        for dataset_name in self.info:
            anatomical_list = list(self.info[dataset_name])
            width = len(anatomical_list) * 5
            ws.merge_cells(start_row=1,
                           start_column=start,
                           end_row=1,
                           end_column=start + width - 1)
            self.cellValue(ws, 1, start, dataset_name)
            a_start = start
            for anatomical_name in anatomical_list:
                ws.merge_cells(start_row=2,
                               start_column=a_start,
                               end_row=2,
                               end_column=a_start + 4)
                self.cellValue(ws, 2, a_start, anatomical_name)
                for _, metric in enumerate(
                    ['Dice', 'HD', 'SymAPD', 'APDgt_pred', 'APDpred_gt']):
                    self.cellValue(ws, 3, a_start + _, metric)
                a_start += 5
            start += width

    def setWorksheet(self, ws: worksheet.worksheet.Worksheet, start_row,
                     network, name, direction):
        alignment = Alignment(vertical='center', horizontal='center')
        self.cellValue(ws, start_row, 1, network)
        self.cellValue(ws, start_row, 2, name)
        metric_name = [
            'BE', 'Jacobian', 'Dice', 'HD', 'SymAPD', 'APDgt_pred',
            'APDpred_gt'
        ]
        metric_format = [
            '%.2f(%.2f)', '%.2f(%.2f)', '%.3f(%.3f)', '%.2f(%.2f)',
            '%.2f(%.2f)', '%.2f(%.2f)', '%.2f(%.2f)'
        ]
        for i in range(3, 10):
            if metric_name[i - 3] in self.mean_by_all[direction]:
                metric_value = self.mean_by_all[direction][metric_name[i - 3]]
                value = metric_format[i - 3] % (metric_value['mean'],
                                                metric_value['std'])
                self.cellValue(ws, start_row, i, value)
        start = 10
        for dataset_name in ['York', 'ACDC', 'MICCAI','M&M']:
            for i in range(7):
                if metric_name[i] in self.mean_by_dataset[dataset_name][
                        direction]:
                    metric_value = self.mean_by_dataset[dataset_name][
                        direction][metric_name[i]]
                    value = metric_format[i] % (metric_value['mean'],
                                                metric_value['std'])
                    self.cellValue(ws, start_row, i + start, value)
            start += 7

        metric_name = ['Dice', 'HD', 'SymAPD', 'APDgt_pred', 'APDpred_gt']
        metric_format = [
            '%.3f(%.3f)', '%.2f(%.2f)', '%.2f(%.2f)', '%.2f(%.2f)',
            '%.2f(%.2f)'
        ]

        for dataset_name in ['York', 'ACDC', 'MICCAI','M&M']:
            anatomical_list = list(self.info[dataset_name])
            for anatomical_name in anatomical_list:
                for i in range(5):
                    metric_value = self.mean_by_anotomical[dataset_name][
                        anatomical_name][direction][metric_name[i]]
                    value = metric_format[i] % (metric_value['mean'],
                                                metric_value['std'])
                    self.cellValue(ws, start_row, i + start, value)
                start += 5

    def saveAsExcel(self, network, name, excel_save_path):
        wb = Workbook()
        wb_name = '%s-%s-%s.xlsx' % (network, name,
                                     time.strftime('%Y%m%d%H%M%S'))
        wb_save_path = os.path.join(excel_save_path, wb_name)

        ws1 = wb.active
        ws1.title = 'ed_to_es'
        self.initWorksheet(ws1)
        self.setWorksheet(ws1, 4, network, name, 'ed_to_es')
        self.autoSetWidth(ws1)

        ws2 = wb.create_sheet(title='es_to_ed')
        self.initWorksheet(ws2)
        self.setWorksheet(ws2, 4, network, name, 'es_to_ed')
        self.autoSetWidth(ws2)

        ws3 = wb.create_sheet(title='undef')
        self.initWorksheet(ws3)
        self.setWorksheet(ws3, 4, 'undef', '', 'undef')
        self.autoSetWidth(ws3)

        wb.save(wb_save_path)
        
        return wb_save_path

    def output(self):
        pprint.pprint(self.mean_by_anotomical)
        pprint.pprint(self.mean_by_dataset)
        pprint.pprint(self.mean_by_all)